import datetime
import re
from copy import copy

import openpyxl
from imomo.managers import ReportDataManager
from imomo.utils.xls.formatters import hydrology_cell_value_formatter
from openpyxl.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from sapphire_backend.imomo import errors


class Tag:
    HEADER_TAG = "HEADER"
    DATA_TAG = "DATA"
    SPLIT = "."
    TAG_START = "{{"
    TAG_END = "}}"
    tag_re = rf"{TAG_START}(.*?){TAG_END}"

    def __init__(
        self,
        tag,
        get_value_method,
        description,
        args=None,
        types=None,
        data=False,
        header=False,
        custom_number_format_fn=None,
    ):
        self.tag = tag
        self.get_value = get_value_method
        self.data = data
        self.header = header
        self.types = "__all__" if types is None else types
        self._description = description
        self.args = args
        self.custom_number_format_fn = custom_number_format_fn

    @property
    def description(self):
        return _(self._description)

    def full_tag(self, special=None):
        if special is not None:
            if special not in (self.HEADER_TAG, self.DATA_TAG):
                raise Exception("Invalid special parameter!")
            special_extension = f"{special}{self.SPLIT}"
        else:
            special_extension = ""
        return f"{self.TAG_START}{special_extension}{self.tag}{self.TAG_END}"

    @property
    def general_tag(self):
        return not (self.header or self.data)

    @property
    def full_header_tag(self):
        return f"{self.TAG_START}{self.tag}{self.TAG_END}"

    def is_valid_type(self, type_name):
        if self.types == "__all__":
            return True
        else:
            return type_name in self.types

    @classmethod
    def decode_tag(cls, tag):
        parts = tag.split(cls.SPLIT)
        return {
            "tag": parts.pop(-1),
            "tag_type": parts.pop(-1) if parts else None,
        }

    def __repr__(self):
        return self.tag

    def __eq__(self, other):
        return other == self.tag

    def __hash__(self):
        return hash(self.tag)


class ReportGenerator:
    sites = []
    session = None
    site_ids = None
    validated = False
    language = "en"

    general_tags = {}

    data_tags = set()
    data_cells = []

    header_tag = None
    header_cell = None

    _ = lambda message: message
    tags = (
        Tag(
            "DATE",
            ReportDataManager.localize_date,
            _("Today's date string."),
        ),
        Tag(
            "DECADE_PERIOD",
            ReportDataManager.decade_period_str,
            _("Decade's period string."),
            types=["decadal_bulletin"],
        ),
        Tag(
            "SITE_BASIN",
            lambda site, **kwargs: site.basin,
            _("Site basin."),
            data=True,
            header=True,
        ),
        Tag(
            "SITE_REGION",
            lambda site, **kwargs: site.region,
            _("Site region."),
            data=True,
            header=True,
        ),
        Tag(
            "SITE_NAME",
            lambda site, **kwargs: site.site_name,
            _("Site name."),
            data=True,
        ),
        Tag(
            "SITE_CODE",
            lambda site, **kwargs: site.site_code,
            _("Site code."),
            data=True,
        ),
        Tag(
            "DISCHARGE_MORNING",
            ReportDataManager.discharge_morning,
            _("Today's morning (8 AM at local time) discharge estimation."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_MORNING_1",
            ReportDataManager.discharge_morning,
            _("Yesterday's morning (8 AM at local time) discharge estimation."),
            data=True,
            args={"date_offset": {"days": -1}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_MORNING_2",
            ReportDataManager.discharge_morning,
            _("Two days ago morning (8 AM at local time) discharge estimation."),
            data=True,
            args={"date_offset": {"days": -2}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_EVENING",
            ReportDataManager.discharge_evening,
            _("Today's evening (8 PM at local time) discharge estimation."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_EVENING_1",
            ReportDataManager.discharge_evening,
            _("Yesterday's evening (8 PM at local time) discharge estimation."),
            data=True,
            args={"date_offset": {"days": -1}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_EVENING_2",
            ReportDataManager.discharge_evening,
            _("Two days ago evening (8 PM at local time) discharge estimation."),
            data=True,
            args={"date_offset": {"days": -2}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DAILY",
            ReportDataManager.discharge_daily,
            _("Today's daily average discharge estimation."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DAILY_1",
            ReportDataManager.discharge_daily,
            _("Yesterday's daily average discharge estimation."),
            data=True,
            args={"date_offset": {"days": -1}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DAILY_2",
            ReportDataManager.discharge_daily,
            _("Two days ago daily average discharge estimation."),
            data=True,
            args={"date_offset": {"days": -2}},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_FIVEDAY",
            ReportDataManager.discharge_fiveday,
            _("Current 5-day period average discharge."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_FIVEDAY_1",
            ReportDataManager.discharge_fiveday,
            _("Previous 5-day period average discharge."),
            data=True,
            args={"period_offset": -1},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE",
            ReportDataManager.discharge_decade,
            _("Current 10-day period average discharge."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE_1",
            ReportDataManager.discharge_decade,
            _("Previous 10-day period average discharge."),
            data=True,
            args={"period_offset": -1},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_MEASUREMENT",
            ReportDataManager.discharge_measurement,
            _("Today's discharge measurement."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_MAX",
            ReportDataManager.discharge_maximum,
            _("Site's discharge maximum recommendation."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE_NORM_C",
            ReportDataManager.get_decade_discharge_norm_cached,
            _(
                "Discharge decade norm on the current day with ignored data newer "
                "than the latest year dividable by 5."
            ),
            data=True,
            args={"cutoff": True},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE_NORM",
            ReportDataManager.get_decade_discharge_norm_cached,
            _("Discharge decade norm on current day."),
            data=True,
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE_1_Y",
            ReportDataManager.get_decade_discharge_norm_cached,
            _("Decade discharge for current decade period on previous year."),
            data=True,
            args={"relative_years_range": (1, 1)},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DECADE_NORM_10_Y",
            ReportDataManager.get_decade_discharge_norm_cached,
            _("Decade discharge average for current decade period for previous 10 years"),
            data=True,
            args={"relative_years_range": (10, None)},
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "WATER_LEVEL_MORNING",
            ReportDataManager.water_level_morning,
            _("Today's morning (8 AM at local time) water level measurement."),
            data=True,
        ),
        Tag(
            "WATER_LEVEL_MORNING_1",
            ReportDataManager.water_level_morning,
            _("Yesterday's morning (8 AM at local time) water level measurement."),
            data=True,
            args={"date_offset": {"days": -1}},
        ),
        Tag(
            "WATER_LEVEL_MORNING_2",
            ReportDataManager.water_level_morning,
            _("Two days ago morning (8 AM at local time) water level measurement."),
            data=True,
            args={"date_offset": {"days": -2}},
        ),
        Tag(
            "WATER_LEVEL_EVENING",
            ReportDataManager.water_level_evening,
            _("Today's evening (8 PM at local time) water level measurement."),
            data=True,
        ),
        Tag(
            "WATER_LEVEL_EVENING_1",
            ReportDataManager.water_level_evening,
            _("Yesterday's evening (8 PM at local time) water level measurement."),
            data=True,
            args={"date_offset": {"days": -1}},
        ),
        Tag(
            "WATER_LEVEL_EVENING_2",
            ReportDataManager.water_level_evening,
            _("Two days ago evening (8 PM at local time) water level measurement."),
            data=True,
            args={"date_offset": {"days": -2}},
        ),
        Tag(
            "WATER_LEVEL_DAILY",
            ReportDataManager.water_level_daily,
            _("Today's water level average."),
            data=True,
        ),
        Tag(
            "WATER_LEVEL_DAILY_1",
            ReportDataManager.water_level_daily,
            _("Yesterday's water level average."),
            data=True,
            args={"date_offset": {"days": -1}},
        ),
        Tag(
            "WATER_LEVEL_DAILY_2",
            ReportDataManager.water_level_daily,
            _("Two days ago water level average."),
            data=True,
            args={"date_offset": {"days": -2}},
        ),
        Tag(
            "WATER_LEVEL_DECADAL_MEASUREMENT",
            ReportDataManager.water_level_measurement,
            _("Water level decadal measurement."),
            data=True,
        ),
        Tag(
            "WATER_LEVEL_MORNING_TREND",
            ReportDataManager.get_trend,
            _("Water level morning (8 AM at local time) trend: today - yesterday value."),
            data=True,
            args={"getter_fn": ReportDataManager.water_level_morning},
        ),
        Tag(
            "WATER_LEVEL_EVENING_TREND",
            ReportDataManager.get_trend,
            _("Water level evening (8 PM at local time) trend: today - yesterday value."),
            data=True,
            args={"getter_fn": ReportDataManager.water_level_evening},
        ),
        Tag(
            "WATER_LEVEL_DAILY_TREND",
            ReportDataManager.get_trend,
            _("Water level daily trend: today - yesterday value."),
            data=True,
            args={"getter_fn": ReportDataManager.water_level_daily},
        ),
        Tag(
            "DISCHARGE_MORNING_TREND",
            ReportDataManager.get_trend,
            _("Discharge morning (8 AM at local time) trend: today - yesterday value."),
            data=True,
            args={
                "getter_fn": ReportDataManager.discharge_morning,
            },
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_EVENING_TREND",
            ReportDataManager.get_trend,
            _("Discharge evening (8 PM at local time) trend: today - yesterday value."),
            data=True,
            args={
                "getter_fn": ReportDataManager.discharge_evening,
            },
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "DISCHARGE_DAILY_TREND",
            ReportDataManager.get_trend,
            _("DISCHARGE daily trend: today - yesterday value."),
            data=True,
            args={
                "getter_fn": ReportDataManager.discharge_daily,
            },
            custom_number_format_fn=hydrology_cell_value_formatter,
        ),
        Tag(
            "ICE_PHENOMENA",
            ReportDataManager.ice_phenomena,
            _("Ice phenomena string value on the requested date."),
            data=True,
        ),
    )

    valid_tags = {tag: tag for tag in tags}

    def __init__(self, template_type_name, template_file):
        self.template_type = template_type_name
        self.workbook = openpyxl.load_workbook(template_file)
        self.sheet = self.workbook.worksheets[0]

    @classmethod
    def get_all_tags(cls, type):
        general_tags = []
        header_tags = []
        data_tags = []

        for tag in cls.tags:
            if tag.types != "__all__" and type not in tag.types:
                continue

            if tag.general_tag:
                general_tags.append(
                    {
                        "tag": tag.full_tag(),
                        "description": tag.description,
                    }
                )
            else:
                if tag.header:
                    header_tags.append(
                        {
                            "tag": tag.full_tag(Tag.HEADER_TAG),
                            "description": tag.description,
                        }
                    )

                if tag.data:
                    data_tags.append(
                        {
                            "tag": tag.full_tag(Tag.DATA_TAG),
                            "description": tag.description,
                        }
                    )

        return {
            "general": general_tags,
            "header": header_tags,
            "data": data_tags,
        }

    @classmethod
    def parse_tags(cls, value):
        try:
            return re.findall(Tag.tag_re, value)
        except TypeError:
            return []

    @classmethod
    def copy_cell_style(cls, source, destination):
        destination.font = copy(source.font)
        destination.border = copy(source.border)
        destination.fill = copy(source.fill)
        destination.number_format = copy(source.number_format)
        destination.protection = copy(source.protection)
        destination.alignment = copy(source.alignment)

    def validate(self):
        self.check_for_tags()
        self.validate_header_and_data_tags()
        self.validated = True

    def validate_tag(self, tag, tag_type):
        try:
            tag_obj = self.valid_tags[tag]
        except KeyError:
            raise errors.ValidationError(_('Invalid tag "{}".').format(tag))

        if tag_type == Tag.HEADER_TAG:
            if not tag_obj.header:
                raise errors.ValidationError(_("Tag {} is not valid HEADER tag.").format(tag))

            if self.header_tag is not None:
                raise errors.ValidationError(_("Can't have multiple HEADER tags."))

        if tag_type == Tag.DATA_TAG:
            if not tag_obj.data:
                raise errors.ValidationError(_("Tag {} is not valid DATA tag.").format(tag))

        return tag_obj

    def validate_header_and_data_tags(self):
        if self.header_tag is None:
            raise errors.ValidationError(_("HEADER tag is required."))

        if not self.data_tags:
            raise errors.ValidationError(_("At least one DATA tag is required."))

        header_row = self.header_cell.row

        data_row = None
        for cell in self.data_cells:
            data_row_ = cell.row
            if data_row is None:
                data_row = data_row_

            if data_row != data_row_:
                raise errors.ValidationError(_("All DATA tags should be in the same row."))

        if data_row - header_row != 1:
            raise errors.ValidationError(_("HEADER tag should be exactly 1 row above DATA tags."))

    def check_for_tags(self):
        self.general_tags = {}
        self.data_tags = set()
        self.data_cells = []
        self.header_tag = None
        self.header_cell = None
        for cell in self.iter_all_cells():
            if cell.value is None:
                continue

            for raw_tag in self.parse_tags(cell.value):
                tag_info = Tag.decode_tag(raw_tag)
                tag_obj = self.validate_tag(**tag_info)

                if tag_info["tag_type"] == Tag.HEADER_TAG:
                    self.header_tag = tag_obj
                    self.header_cell = cell

                elif tag_info["tag_type"] == Tag.DATA_TAG:
                    self.data_tags.add(tag_obj)
                    self.data_cells.append(cell)
                else:
                    if tag_obj not in self.general_tags:
                        self.general_tags[tag_obj] = []

                    self.general_tags[tag_obj].append(cell)

    def iter_all_cells(self):
        for row in self.sheet.rows:
            yield from row

    def copy_range(self, range_start, range_end, dest_ranges):
        source_rows = self.sheet.rows[range_start[0] - 1 : range_end[0]]

        for dest_range_start in dest_ranges:
            dest_row = dest_range_start[0]
            dest_col = dest_range_start[1]

            for row in source_rows:
                cells = row[range_start[1] - 1 : range_end[1]]
                for cell in cells:
                    self.move_cell(
                        source_cell=cell,
                        dest_row=dest_row,
                        dest_col=dest_col,
                        preserve_original=True,
                        move_merged=True,
                    )
                    dest_col += 1

                dest_row += 1

    @staticmethod
    def replace_value(cell, value, full_tag):
        if cell.value == full_tag:
            cell.value = value
        else:
            try:
                cell.value = cell.value.replace(full_tag, unicode(value))
            except AttributeError:
                pass

    @staticmethod
    def move_cell(source_cell, dest_row, dest_col, preserve_original=False, move_merged=False):
        worksheet = source_cell.parent

        if preserve_original:
            cell_to_move = Cell(worksheet)
            cell_to_move.value = source_cell.value
            cell_to_move.font = source_cell.font
            cell_to_move.border = source_cell.border
            cell_to_move.fill = source_cell.fill
            cell_to_move.number_format = copy(source_cell.number_format)
            cell_to_move.protection = source_cell.protection
            cell_to_move.alignment = source_cell.alignment

        else:
            cell_to_move = source_cell

        # cell_to_move.number_format = deepcopy(source_cell.number_format)
        # cell_to_move._style = copy(source_cell._style)
        source_address = (source_cell.row, source_cell.col_idx)
        dest_address = (dest_row, dest_col)

        cell_to_move.row = dest_row
        cell_to_move.col_idx = dest_col
        worksheet._cells[dest_address] = cell_to_move

        if cell_to_move.data_type == "f":
            adr = "{}{}"
            moved_cell = worksheet._cells[dest_address]
            moved_cell.value = Translator(
                cell_to_move.value, adr.format(get_column_letter(source_address[1]), source_address[0])
            ).translate_formula(adr.format(get_column_letter(dest_col), dest_row))
            worksheet._cells[dest_address] = moved_cell

        merges_in_range = []
        if move_merged:
            for range_ in worksheet.merged_cell_ranges:
                merges_in_range.append(range_)
                start_cell_, end_cell_ = range_.split(":")
                if source_cell == worksheet[start_cell_]:
                    source_end_cell = worksheet[end_cell_]
                    row_diff = dest_row - source_cell.row
                    col_diff = dest_col - source_cell.col_idx

                    row_idx = source_end_cell.row + row_diff - 1
                    col_idx = source_end_cell.col_idx + col_diff - 1
                    # end_cell = worksheet.rows[row_idx][col_idx]
                    end_cell = worksheet._cells[(row_idx + 1, col_idx + 1)]

                    worksheet.merged_cell_ranges.append(
                        "{}{}:{}{}".format(
                            cell_to_move.column,
                            cell_to_move.row,
                            end_cell.column,
                            end_cell.row,
                        )
                    )

        if not preserve_original:
            for range_ in merges_in_range:
                worksheet.merged_cell_ranges.remove(range_)

            del worksheet._cells[source_address]

        return cell_to_move

    def generate_report(self, filename, session, sites, date=None, language="en"):
        if not self.validated:
            raise Exception("Template is not validated!")

        date = date or datetime.datetime.now()

        self.session = session
        self.sites = sites
        self.language = language

        sheet = self.sheet

        header_row = self.header_cell.row
        header_col = self.header_cell.col_idx
        data_first_row = header_row + 1

        context = {"sites": sites}

        # group sites by header
        grouped_sites = {}
        for site in self.sites:
            header_value = self.header_tag.get_value(
                site=site,
                date=date,
                session=session,
                context=context,
            )
            if header_value not in grouped_sites:
                grouped_sites[header_value] = []

            grouped_sites[header_value].append(site)

        # insert empty rows
        number_of_columns = len(self.sites) + len(grouped_sites) - 2
        sheet.insert_rows(data_first_row, number_of_columns, max_column=25)

        # c/p header and data rows (values and style)
        current_row = header_row

        data_tags_dest_ranges = []
        header_tags_dest_ranges = []
        for group_header, sites in grouped_sites.iteritems():
            # header ranges
            if current_row != header_row:
                header_tags_dest_ranges.append((current_row, header_col))

            current_row += 1
            for x in sites:
                # data ranges
                if current_row != data_first_row:
                    data_tags_dest_ranges.append((current_row, 1))
                current_row += 1

        self.copy_range(
            (header_row, header_col),
            (header_row, header_col),
            header_tags_dest_ranges,
        )

        self.copy_range(
            (data_first_row, 1),
            (data_first_row, 25),
            data_tags_dest_ranges,
        )

        current_row = header_row
        for group_header, sites in grouped_sites.iteritems():
            # replace header tags with header data
            dest_cell = sheet._cells[current_row, header_col]
            self.replace_value(dest_cell, group_header, self.header_tag.full_tag(Tag.HEADER_TAG))
            current_row += 1

            for site in sites:
                for data_cell in self.data_cells:
                    # replace data tags with site related data
                    dest_cell = sheet._cells[current_row, data_cell.col_idx]
                    for raw_tag in self.parse_tags(dest_cell.value):
                        tag_info = Tag.decode_tag(raw_tag)
                        tag_obj = self.valid_tags[tag_info["tag"]]
                        tag_args = tag_obj.args if tag_obj.args else {}
                        value = tag_obj.get_value(site=site, date=date, session=session, context=context, **tag_args)
                        if tag_obj.custom_number_format_fn:
                            num_format_str, value = tag_obj.custom_number_format_fn(value)
                            dest_cell.number_format = num_format_str

                        self.replace_value(dest_cell, value, tag_obj.full_tag(Tag.DATA_TAG))

                current_row += 1

        for tag, cells in self.general_tags.iteritems():
            # replace general tags with general data
            tag_value = tag.get_value(date=date, session=session, context=context, language=language)

            for cell in cells:
                self.replace_value(cell, tag_value, tag.full_tag())

        self.workbook.save(filename)
