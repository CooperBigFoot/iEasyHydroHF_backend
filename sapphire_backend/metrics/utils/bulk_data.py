import pandas as pd
from openpyxl.utils import get_column_letter

from sapphire_backend.metrics.models import BulkDataHydroAuto, BulkDataHydroManual, BulkDataMeteo, BulkDataVirtual
from sapphire_backend.stations.models import HydrologicalStation, MeteorologicalStation, VirtualStation


def add_sheet_column_styling(writer: pd.ExcelWriter, sheet_name: str, sheet_headers: list):
    worksheet = writer.sheets[sheet_name]

    column_letter = get_column_letter(1)
    worksheet.column_dimensions[column_letter].width = 20
    for col_num in range(2, len(sheet_headers) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 18
    worksheet.row_dimensions[1].height = 45


def transform_timestamp_local_format(df: pd.DataFrame) -> pd.DataFrame:
    df["timestamp_local"] = pd.to_datetime(df["timestamp_local"]).dt.strftime("%d.%m.%Y. %H:%M:%S")
    return df


def transform_decimal_columns(df: pd.DataFrame, decimal_columns: list[str]) -> pd.DataFrame:
    for col in decimal_columns:
        df[col] = df[col].apply(lambda x: float(x) if x is not None else "")

    return df


def transform_code_value_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    def remove_trailing_zeros(number_str: str):
        code, value = number_str.split(":")
        if "." in value:
            value = value.rstrip("0").rstrip(".")
        return f"{code}:{value}"

    for col in columns:
        df[col] = df[col].apply(lambda x: remove_trailing_zeros(x) if x is not None else "")

    return df


def write_bulk_data_hydro_manual_sheets(
    writer: pd.ExcelWriter, station_uuids: list[str], org_uuid: str, all: bool = False
) -> pd.DataFrame:
    if not all and len(station_uuids) == 0:
        return
    excel_sheet_headers = [
        "\n\nDate / Unit",
        "Water level\ndaily\n(cm)",
        "Water level\ndaily average\n(cm)",
        "Discharge\nmeasurement\n(m^3/s)",
        "Discharge\ndaily\n(m^3/s)",
        "Free river area\n\n(m^2)",
        "Decade\ndischarge\n(m^3/s)",
        "Discharge\ndaily average\n(m^3/s)",
        "Ice phenomena\n\n(code:intensity)",
        "Water level\nmeasurement\n(cm)",
        "Fiveday\ndischarge\n(m^3/s)",
        "Air\ntemperature\n(°C)",
        "Water\ntemperature\n(°C)",
        "Daily\nprecipitation\n(code:mm)",
    ]

    df_columns = [
        "station",
        "timestamp_local",
        "water_level_daily",
        "water_level_daily_average",
        "discharge_measurement",
        "discharge_daily",
        "free_river_area",
        "decade_discharge",
        "discharge_daily_average",
        "ice_phenomena",
        "water_level_measurement",
        "fiveday_discharge",
        "air_temperature",
        "water_temperature",
        "precipitation_daily",
    ]

    if not all:
        hydro_stations = HydrologicalStation.objects.filter(
            uuid__in=station_uuids, station_type="M", site__organization__uuid=org_uuid
        ).values("id", "station_code")
    else:
        hydro_stations = HydrologicalStation.objects.filter(
            station_type="M", site__organization__uuid=org_uuid
        ).values("id", "station_code")

    hydro_station_ids = tuple(hydro_stations.values_list("id", flat=True))
    hydro_queryset = (
        BulkDataHydroManual.objects.filter(station_id__in=hydro_station_ids)
        .values_list(*df_columns)
        .order_by("timestamp_local")
    )

    df = pd.DataFrame.from_records(hydro_queryset, columns=df_columns)

    df = transform_timestamp_local_format(df)
    df = transform_decimal_columns(
        df,
        decimal_columns=[
            "water_level_daily",
            "water_level_daily_average",
            "discharge_measurement",
            "discharge_daily",
            "free_river_area",
            "decade_discharge",
            "discharge_daily_average",
            "water_level_measurement",
            "air_temperature",
            "water_temperature",
            "fiveday_discharge",
        ],
    )
    df = transform_code_value_columns(df, columns=["ice_phenomena", "precipitation_daily"])

    grouped = df.groupby("station")

    for station, group in grouped:
        group = group.drop(columns=["station"])
        group.columns = excel_sheet_headers
        station_code = hydro_stations.get(id=station)["station_code"]
        sheet_name = f"{station_code} (manual)"

        group.to_excel(writer, sheet_name=sheet_name, index=False)
        add_sheet_column_styling(writer, sheet_name, excel_sheet_headers)


def write_bulk_data_hydro_auto_sheets(
    writer: pd.ExcelWriter, station_uuids: list[str], org_uuid: str, all: bool = False
) -> pd.DataFrame:
    if not all and len(station_uuids) == 0:
        return
    excel_sheet_headers = [
        "\n\nDate / Unit",
        "Water level\ndaily minimum\n(cm)",
        "Water level\ndaily average\n(cm)",
        "Water level\ndaily maximum\n(cm)",
        "Air temperature\nminimum\n(°C)",
        "Air temperature\naverage\n(°C)",
        "Air temperature\nmaximum\n(°C)",
        "Water temperature\nminimum\n(°C)",
        "Water temperature\naverage\n(°C)",
        "Water temperature\nmaximum\n(°C)",
        "Discharge\ndaily average\n(m^3/s)",
        "Fiveday\ndischarge\n(m^3/s)",
        "Decade\ndischarge\n(m^3/s)",
    ]

    df_columns = [
        "station",
        "timestamp_local",
        "water_level_daily_min",
        "water_level_daily_average",
        "water_level_daily_max",
        "air_temperature_min",
        "air_temperature_average",
        "air_temperature_max",
        "water_temperature_min",
        "water_temperature_average",
        "water_temperature_max",
        "discharge_daily_average",
        "fiveday_discharge",
        "decade_discharge",
    ]

    if not all:
        hydro_stations = HydrologicalStation.objects.filter(
            uuid__in=station_uuids, site__organization__uuid=org_uuid, station_type="A"
        ).values("id", "station_code")
    else:
        hydro_stations = HydrologicalStation.objects.filter(
            site__organization__uuid=org_uuid, station_type="A"
        ).values("id", "station_code")

    hydro_station_ids = tuple(hydro_stations.values_list("id", flat=True))
    hydro_queryset = (
        BulkDataHydroAuto.objects.filter(station_id__in=hydro_station_ids)
        .values_list(*df_columns)
        .order_by("timestamp_local")
    )

    df = pd.DataFrame.from_records(hydro_queryset, columns=df_columns)

    df = transform_timestamp_local_format(df)
    df = transform_decimal_columns(
        df,
        decimal_columns=[
            "water_level_daily_min",
            "water_level_daily_average",
            "water_level_daily_max",
            "air_temperature_min",
            "air_temperature_average",
            "air_temperature_max",
            "water_temperature_min",
            "water_temperature_average",
            "water_temperature_max",
            "discharge_daily_average",
            "fiveday_discharge",
            "decade_discharge",
        ],
    )

    grouped = df.groupby("station")

    for station, group in grouped:
        group = group.drop(columns=["station"])
        group.columns = excel_sheet_headers
        station_code = hydro_stations.get(id=station)["station_code"]
        sheet_name = f"{station_code} (auto)"

        group.to_excel(writer, sheet_name=sheet_name, index=False)
        add_sheet_column_styling(writer, sheet_name, excel_sheet_headers)


def write_bulk_data_meteo_sheets(
    writer: pd.ExcelWriter, station_uuids: list[str], org_uuid: str, all: bool = False
) -> pd.DataFrame:
    if not all and len(station_uuids) == 0:
        return
    excel_sheet_headers = [
        "\n\nDate / Unit",
        "Precipitation\ndecade\n(mm/day)",
        "Precipitation\nmonthly\n(mm/day)",
        "Air temperature\ndecade\n(°C)",
        "Air temperature\nmonthly\n(°C)",
    ]

    df_columns = [
        "station",
        "timestamp_local",
        "precipitation_decade_average",
        "precipitation_month_average",
        "air_temperature_decade_average",
        "air_temperature_month_average",
    ]

    if not all:
        meteo_stations = MeteorologicalStation.objects.filter(
            uuid__in=station_uuids, site__organization__uuid=org_uuid
        ).values("id", "station_code")
    else:
        meteo_stations = MeteorologicalStation.objects.filter(site__organization__uuid=org_uuid).values(
            "id", "station_code"
        )

    meteo_station_ids = tuple(meteo_stations.values_list("id", flat=True))
    meteo_queryset = (
        BulkDataMeteo.objects.filter(station_id__in=meteo_station_ids)
        .values_list(*df_columns)
        .order_by("timestamp_local")
    )
    df = pd.DataFrame.from_records(meteo_queryset, columns=df_columns)

    df = transform_timestamp_local_format(df)
    df = transform_decimal_columns(
        df,
        decimal_columns=[
            "precipitation_decade_average",
            "precipitation_month_average",
            "air_temperature_decade_average",
            "air_temperature_month_average",
        ],
    )

    grouped = df.groupby("station")

    for station, group in grouped:
        group = group.drop(columns=["station"])
        group.columns = excel_sheet_headers
        station_code = meteo_stations.get(id=station)["station_code"]
        sheet_name = f"{station_code} (meteo)"

        group.to_excel(writer, sheet_name=sheet_name, index=False)
        add_sheet_column_styling(writer, sheet_name, excel_sheet_headers)


def write_bulk_data_virtual_sheets(
    writer: pd.ExcelWriter, station_uuids: list[str], org_uuid: str, all: bool = False
) -> pd.DataFrame:
    if not all and len(station_uuids) == 0:
        return
    excel_sheet_headers = [
        "\n\nDate / Unit",
        "Discharge\ndaily\n(m^3/s)",
        "Decade\ndischarge\n(m^3/s)",
        "Discharge\ndaily average\n(m^3/s)",
        "Fiveday\ndischarge\n(m^3/s)",
    ]

    df_columns = [
        "station",
        "timestamp_local",
        "discharge_daily",
        "decade_discharge",
        "discharge_daily_average",
        "fiveday_discharge",
    ]

    if not all:
        virtual_stations = VirtualStation.objects.filter(uuid__in=station_uuids, organization__uuid=org_uuid).values(
            "id", "station_code"
        )
    else:
        virtual_stations = VirtualStation.objects.filter(organization__uuid=org_uuid).values("id", "station_code")

    virtual_station_ids = tuple(virtual_stations.values_list("id", flat=True))

    virtual_queryset = (
        BulkDataVirtual.objects.filter(station_id__in=virtual_station_ids)
        .values_list(*df_columns)
        .order_by("timestamp_local")
    )

    df = pd.DataFrame.from_records(virtual_queryset, columns=df_columns)

    df = transform_timestamp_local_format(df)
    df = transform_decimal_columns(
        df,
        decimal_columns=[
            "discharge_daily",
            "decade_discharge",
            "discharge_daily_average",
            "fiveday_discharge",
        ],
    )

    grouped = df.groupby("station")

    for station, group in grouped:
        group = group.drop(columns=["station"])
        group.columns = excel_sheet_headers
        station_code = virtual_stations.get(id=station)["station_code"]
        sheet_name = f"{station_code} (virtual)"

        group.to_excel(writer, sheet_name=sheet_name, index=False)
        add_sheet_column_styling(writer, sheet_name, excel_sheet_headers)
